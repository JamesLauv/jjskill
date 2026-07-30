"""
Excel多Sheet生成脚本
从结构化JSON输入生成带样式的多Sheet Excel文件。

用法: python gen_excel.py <input.json> <output.xlsx>

JSON输入格式:
{
  "title": "文档标题",
  "subtitle": "副标题",
  "sheets": [
    {
      "name": "Sheet名称",
      "tab_color": "FF9800",
      "headers": ["列1", "列2", ...],
      "widths": [14, 20, ...],
      "header_fill": "283593",
      "rows": [
        ["数据1", "数据2", ...],
        ...
      ],
      "row_height": 35,
      "merge_columns": [[0, 5], [6, 8]],  // 可选：合并行范围 [start_row, end_row]
      "freeze_pane": "A5"  // 可选：冻结窗格
    }
  ],
  "extra_columns": [  // 可选：在第一个Sheet追加列
    {
      "col_index": 8,
      "header": "列名",
      "header_fill": "283593",
      "width": 40,
      "data": {
        "5": {"value": "内容", "font_color": "D32F2F", "font_size": 9}
      }
    }
  ]
}
"""
import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# 通用样式
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name="微软雅黑", size=10, color="333333")
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True, color="1A237E")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)


def create_sheet(wb, sheet_config, is_first=False):
    """创建一个Sheet"""
    ws = wb.create_sheet(sheet_config["name"])
    if "tab_color" in sheet_config:
        ws.sheet_properties.tabColor = sheet_config["tab_color"]

    headers = sheet_config["headers"]
    widths = sheet_config.get("widths", [16] * len(headers))
    header_fill_color = sheet_config.get("header_fill", "283593")
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    row_height = sheet_config.get("row_height", 35)

    # 写标题行（如果有）
    start_row = 1
    if "title" in sheet_config:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(row=1, column=1, value=sheet_config["title"]).font = TITLE_FONT
        ws.cell(row=1, column=1).alignment = CENTER_ALIGN
        ws.row_dimensions[1].height = 38
        start_row = 3

    # 写表头
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=start_row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[start_row].height = 28

    # 写数据行
    data_start = start_row + 1
    for ri, row_data in enumerate(sheet_config.get("rows", []), data_start):
        for ci, v in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=v)
            cell.font = NORMAL_FONT
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER
        ws.row_dimensions[ri].height = row_height

    # 合并指定列的单元格
    if "merge_columns" in sheet_config:
        for col_idx, (start, end) in enumerate(sheet_config["merge_columns"], 1):
            if start != end:
                ws.merge_cells(
                    start_row=data_start + start, start_column=1,
                    end_row=data_start + end, end_column=1
                )
                ws.cell(row=data_start + start, column=1).alignment = CENTER_ALIGN

    # 冻结窗格
    if "freeze_pane" in sheet_config:
        ws.freeze_panes = sheet_config["freeze_pane"]
    else:
        ws.freeze_panes = "A%d" % (start_row + 1)

    return ws


def add_extra_columns(ws, extra_cols, data_start_row):
    """追加额外列到已有Sheet"""
    for col_info in extra_cols:
        col_idx = col_info["col_index"]
        header = col_info["header"]
        header_fill = PatternFill(
            start_color=col_info.get("header_fill", "283593"),
            end_color=col_info.get("header_fill", "283593"),
            fill_type="solid"
        )

        cell = ws.cell(row=data_start_row - 1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_info.get("width", 30)

        for row_str, val_info in col_info.get("data", {}).items():
            row_num = int(row_str)
            cell = ws.cell(row=row_num, column=col_idx, value=val_info["value"])
            cell.font = Font(
                name="微软雅黑",
                size=val_info.get("font_size", 10),
                color=val_info.get("font_color", "333333")
            )
            cell.alignment = LEFT_ALIGN
            cell.border = THIN_BORDER


def main():
    if len(sys.argv) < 3:
        print("用法: python gen_excel.py <input.json> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        config = json.load(f)

    wb = Workbook()
    # 删除默认Sheet
    wb.remove(wb.active)

    # 创建所有Sheet
    first_ws = None
    for si, sheet_config in enumerate(config.get("sheets", [])):
        ws = create_sheet(wb, sheet_config, is_first=(si == 0))
        if si == 0:
            first_ws = ws

    # 追加额外列
    if "extra_columns" in config and first_ws:
        # 计算数据起始行
        first_sheet = config["sheets"][0]
        data_start = 3 if "title" in first_sheet else 2
        add_extra_columns(first_ws, config["extra_columns"], data_start)

    wb.save(output_path)
    print("Done: %s" % output_path)


if __name__ == "__main__":
    main()
